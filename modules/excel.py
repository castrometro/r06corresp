import pandas as pd
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

def status_modulo_excel():
    print("Modulo de manejo de archivos excel funcionando exitosamente")
    return True

def excel_to_df(path, header, columns):
    if header != None and columns != None:
        df = pd.read_excel(path, engine='openpyxl', header= header, usecols = columns)
    
    elif header != None and columns == None:
        df = pd.read_excel(path, engine='openpyxl', header= header)
    
    elif header == None and columns != None:
        df = pd.read_excel(path, engine='openpyxl', usecols = columns)

    else:
        df = pd.read_excel(path, engine='openpyxl')
    return df

def xls_to_df(path):
    df = pd.read_excel(path, engine = 'xlrd')
    return df

def list_to_df(lista, nombre_columna):
    df = pd.DataFrame(lista, columns=[nombre_columna])
    return df

def agrupar_df(df, lista_columnas_agrupar, lista_columnas_sumar):
    df = df.groupby(lista_columnas_agrupar, as_index=False)[lista_columnas_sumar].sum()
    return df

def merge_df(df1, df2, columna_comun):
    df_outer = pd.merge(df1, df2, on=columna_comun, how='outer', indicator=True)
    unique_df1 = df_outer[df_outer['_merge'] == 'left_only'][columna_comun].tolist()
    unique_df2 = df_outer[df_outer['_merge'] == 'right_only'][columna_comun].tolist()
    df = pd.merge(df1, df2, on=columna_comun, how='inner')
    return df,unique_df1,unique_df2

def operatoria_entre_columnas(df, columna1, columna2, operacion, resultado):
    if operacion == '+':
        df[resultado] = df[columna1] + df[columna2]
    elif operacion == '-':
        df[resultado] = df[columna1] - df[columna2]
    elif operacion == '*':
        df[resultado] = df[columna1] * df[columna2]
    elif operacion == '/':
        df[resultado] = df[columna1] / df[columna2]
    return df

def df_to_excel(df, nombre, path):
    full_path = os.path.join(path, nombre + '.xlsx')
    df.to_excel(full_path, index=False)
    return full_path

def formatear_columna_df(ruta_excel, nombre_hoja, nombre_columna, formato):
    wb = load_workbook(ruta_excel)
    if nombre_hoja is None:
        ws = wb.active
    else:
        ws = wb[nombre_hoja]
    column_index = None
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            if cell.value == nombre_columna:
                column_index = cell.column_letter
                break
        if column_index is not None:
            break

    if column_index is None:
        raise ValueError(f"No se encontró la columna con el encabezado '{nombre_columna}' en la hoja de trabajo '{nombre_hoja}'.")

    # Aplicar el formato según el tipo especificado
    if formato == 'porcentaje':
        format_string = '0.00%'
    elif formato == 'fecha':
        format_string = 'dd/mm/yyyy'
    elif formato == 'moneda':
        format_string = '$#,##0.00'
    elif formato == 'texto':
        format_string = 'General'
    elif formato == 'cientifico':
        format_string = '0.00E+00'
    elif formato == 'fecha_hora':
        format_string = 'dd/mm/yyyy hh:mm:ss'
    elif formato == 'hora':
        format_string = 'hh:mm:ss'
    elif formato == 'contabilidad':
        format_string = '[$$-es-CL]#,##0.00'
    elif formato == 'numero':
        format_string = '#,##0.00'
    else:
        raise ValueError("Formato no válido. Los formatos disponibles son 'currency' y 'number'.")

    # Obtener todas las celdas de la columna identificada por column_index
    column_cells = ws[f"{column_index}"]

    # Aplicar el formato a cada celda en la columna encontrada
    for cell in column_cells:
        cell.number_format = format_string

    # Guardar los cambios en el archivo Excel
    wb.save(ruta_excel)

def filtrar_filas_por_valor(df, columna, valor):
    """
    Filtra el DataFrame para mantener solo las filas donde la columna especificada tiene el valor dado.
    
    Parámetros:
        df (pandas.DataFrame): El DataFrame a filtrar.
        columna (str): El nombre de la columna en la que buscar el valor.
        valor (any): El valor que las filas deben tener en la columna especificada para ser conservadas.
        
    Retorna:
        pandas.DataFrame: Un nuevo DataFrame con las filas filtradas.
    """
    # Filtrar el DataFrame
    df_filtrado = df[df[columna] == valor]

    return df_filtrado

import pandas as pd

def filtrar_filas_por_valores(df, columna, valores):
    """
    Filtra el DataFrame para mantener solo las filas donde la columna especificada contiene alguno de los valores dados.
    
    Parámetros:
        df (pandas.DataFrame): El DataFrame a filtrar.
        columna (str): El nombre de la columna en la que buscar los valores.
        valores (list): Una lista de valores que las filas deben tener en la columna especificada para ser conservadas.
        
    Retorna:
        pandas.DataFrame: Un nuevo DataFrame con las filas filtradas.
    """
    # Filtrar el DataFrame usando isin para verificar múltiples valores en la columna
    df_filtrado = df[df[columna].isin(valores)]
    
    return df_filtrado



def suma_columna_excel(excel_path, sheet_name, column_name):
    # Load the Excel workbook
    wb = pd.ExcelFile(excel_path)
    
    # Use the first sheet if sheet_name is None
    if sheet_name is None:
        sheet_name = wb.sheet_names[0]
    
    # Load the specified Excel sheet
    df = pd.read_excel(wb, sheet_name=sheet_name)
    
    # Check if the column exists in the DataFrame
    if column_name not in df.columns:
        raise ValueError(f"The column '{column_name}' does not exist in the sheet '{sheet_name}'.")

    # Load workbook using openpyxl for editing
    wb_openpyxl = openpyxl.load_workbook(excel_path)
    ws = wb_openpyxl[sheet_name]

    # Define the cell location to place the SUM formula (after the last row of the column)
    last_row_index = len(df) + 2  # Excel indices start from 1, and include header row, add one more for next empty row

    # Find the Excel column letter corresponding to the column name
    column_letter = openpyxl.utils.cell.get_column_letter((df.columns.get_loc(column_name) + 1))

    # Define the formula
    formula = f"=SUM({column_letter}2:{column_letter}{last_row_index - 1})"

    # Assign formula to the cell right below the last data cell in the column
    ws[f"{column_letter}{last_row_index}"].value = formula
    
    # Save the modified workbook
    wb_openpyxl.save(excel_path)
    wb_openpyxl.close()

def sumar_columna_df(df, nombre_columna):
    suma = df[nombre_columna].sum()
    return suma

def buscar_df1_df2_columna(df1, df2, columna,nombre_columna_df1,nombre_columna_df2):
    df1[nombre_columna_df1] = df1[columna].isin(df2[columna])
    df2[nombre_columna_df2] = df2[columna].isin(df1[columna])
    return df1,df2

def buscar_fila_df(df, columna, valor):
    df = df.loc[df[columna] == valor]
    return df

def buscar_filas_df_contiene(df, columna, lista):
    df_temp = df[df[columna].isin(lista)]
    return df_temp


#validar datos por ejemplo formato de ruts.
#19688147-5
#1934
#eliminar guion y dp. Espacios, etc...


def ajustar_ancho_columnas(ruta_excel, ancho_pixeles, num_hoja=None):
    try:
        # Cargamos el libro de Excel
        libro = openpyxl.load_workbook(ruta_excel)
        
        if num_hoja is None:
            hojas = libro.sheetnames
        else:
            hojas = [libro.sheetnames[num_hoja - 1]]
        
        for hoja_nombre in hojas:
            hoja = libro[hoja_nombre]
            # Ajustamos el ancho de todas las columnas
            for columna in hoja.columns:
                columna_letra = columna[0].column_letter
                hoja.column_dimensions[columna_letra].width = ancho_pixeles

        # Guardamos los cambios en el archivo
        libro.save(ruta_excel)
        libro.close()
        if num_hoja is None:
            print(f"Ancho de todas las columnas ajustado a {ancho_pixeles} píxeles en todas las hojas del archivo {ruta_excel}.")
        else:
            print(f"Ancho de todas las columnas ajustado a {ancho_pixeles} píxeles en la hoja {num_hoja} del archivo {ruta_excel}.")
    except Exception as e:
        print(f"Error al ajustar el ancho de las columnas: {e}")

